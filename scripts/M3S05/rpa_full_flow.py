"""
[M3S05] - Ex. 4 - Fluxo RPA Completo com Arquivos

Script que executa um fluxo completo de RPA utilizando Selenium:

  FASE 1 — Ler planilha de URLs de entrada (pandas)
  FASE 2 — Coletar dados de cada URL com Selenium
  FASE 3 — Consolidar dados em DataFrames
  FASE 4 — Salvar em planilha formatada (openpyxl)
  FASE 5 — Renomear arquivo de saída com timestamp

Requisitos cumpridos:
  ✓ Coleta de 2+ links armazenados em planilha
  ✓ Coleta de informações de cada link via Selenium
  ✓ Armazenamento em planilha formatada
  ✓ Renomeação do arquivo de saída

Pré-requisitos:
    - Google Chrome instalado
    - pip install selenium pandas openpyxl

Uso:
    python scripts/M3S05/rpa_full_flow.py
"""

import random
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

# ── Caminhos ─────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parents[2]
PLANILHA_URLS = ROOT / "data" / "input" / "xlsx" / "m3s05_urls_coleta.xlsx"
DIR_SAIDA = ROOT / "data" / "output" / "xlsx"
ARQUIVO_SAIDA_BRUTO = DIR_SAIDA / "coleta_rpa_bruto.xlsx"


# ── Configuração do Selenium ────────────────────────────────────────────────
def criar_driver() -> webdriver.Chrome:
    """Cria e retorna um driver Chrome com configuração padrão do curso."""
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument(
        "user-agent=Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
    )
    return webdriver.Chrome(options=options)


# ── Funções de scraping por tipo ─────────────────────────────────────────────
def scrape_paises(driver: webdriver.Chrome, url: str) -> pd.DataFrame:
    """Coleta dados de países do scrapethissite.com/pages/simple/.

    Args:
        driver: instância do Selenium WebDriver.
        url: URL da página de países.

    Returns:
        DataFrame com colunas [pais, capital, populacao, area_km2].
    """
    driver.get(url)
    WebDriverWait(driver, 10).until(
        EC.presence_of_all_elements_located(
            (By.CSS_SELECTOR, "div.col-md-4.country")
        )
    )

    elementos = driver.find_elements(By.CSS_SELECTOR, "div.col-md-4.country")
    dados = []

    for elem in elementos:
        nome = elem.find_element(By.CSS_SELECTOR, "h3.country-name").text.strip()
        capital = elem.find_element(
            By.CSS_SELECTOR, "span.country-capital"
        ).text.strip()
        populacao = elem.find_element(
            By.CSS_SELECTOR, "span.country-population"
        ).text.strip()
        area = elem.find_element(
            By.CSS_SELECTOR, "span.country-area"
        ).text.strip()

        dados.append({
            "pais": nome,
            "capital": capital,
            "populacao": int(populacao) if populacao else 0,
            "area_km2": float(area) if area else 0.0,
        })

    return pd.DataFrame(dados)


def scrape_hockey(driver: webdriver.Chrome, url: str) -> pd.DataFrame:
    """Coleta dados de times de hockey do scrapethissite.com/pages/forms/.

    Coleta dados da primeira página de resultados (com paginação).

    Args:
        driver: instância do Selenium WebDriver.
        url: URL da página de hockey teams.

    Returns:
        DataFrame com colunas [time, ano, vitorias, derrotas, gols_feitos,
        gols_sofridos, diferenca_gols].
    """
    driver.get(url)
    WebDriverWait(driver, 10).until(
        EC.presence_of_all_elements_located(
            (By.CSS_SELECTOR, "tr.team")
        )
    )

    linhas = driver.find_elements(By.CSS_SELECTOR, "tr.team")
    dados = []

    for linha in linhas:
        colunas = linha.find_elements(By.TAG_NAME, "td")
        if len(colunas) >= 9:
            def safe_int(txt):
                txt = txt.strip()
                return int(txt) if txt else 0

            def safe_float(txt):
                txt = txt.strip()
                return float(txt) if txt else 0.0

            dados.append({
                "time": colunas[0].text.strip(),
                "ano": safe_int(colunas[1].text),
                "vitorias": safe_int(colunas[2].text),
                "derrotas": safe_int(colunas[3].text),
                "derrotas_ot": safe_int(colunas[4].text),
                "pct_vitorias": safe_float(colunas[5].text),
                "gols_feitos": safe_int(colunas[6].text),
                "gols_sofridos": safe_int(colunas[7].text),
                "diferenca_gols": safe_int(colunas[8].text),
            })

    df = pd.DataFrame(dados)
    return df


SCRAPERS = {
    "paises": scrape_paises,
    "hockey": scrape_hockey,
}


# ── Funções de formatação da planilha ────────────────────────────────────────
def formatar_planilha(wb: Workbook) -> None:
    """Aplica formatação profissional a todas as abas da planilha.

    - Header: fundo azul escuro (#1F4E79), texto branco, negrito
    - Dados: bordas finas, alinhamento centralizado para números
    - Colunas: largura ajustada automaticamente

    Args:
        wb: Workbook do openpyxl para formatar.
    """
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    borda_fina = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    data_font = Font(name="Calibri", size=10)

    for ws in wb.worksheets:
        # Formatar header (primeira linha)
        for col_idx in range(1, ws.max_column + 1):
            celula = ws.cell(row=1, column=col_idx)
            celula.fill = header_fill
            celula.font = header_font
            celula.alignment = header_alignment
            celula.border = borda_fina

        # Formatar dados
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                celula = ws.cell(row=row_idx, column=col_idx)
                celula.font = data_font
                celula.border = borda_fina

                # Números: alinhamento à direita
                if isinstance(celula.value, (int, float)):
                    celula.alignment = Alignment(horizontal="right")
                else:
                    celula.alignment = Alignment(horizontal="left")

        # Auto-ajustar largura das colunas
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            for row_idx in range(1, ws.max_row + 1):
                celula = ws.cell(row=row_idx, column=col_idx)
                if celula.value:
                    max_length = max(max_length, len(str(celula.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 4, 50)


def renomear_arquivo(caminho: Path) -> Path:
    """Renomeia o arquivo de saída com timestamp.

    Padrão: {YYYYMMDD_HHMMSS}_coleta_rpa_m3s05.xlsx

    Args:
        caminho: caminho do arquivo original.

    Returns:
        Novo caminho após renomeação.
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    novo_nome = f"{timestamp}_coleta_rpa_m3s05{caminho.suffix}"
    novo_caminho = caminho.parent / novo_nome
    caminho.rename(novo_caminho)
    return novo_caminho


# ═══════════════════════════════════════════════════════════════
# FLUXO PRINCIPAL
# ═══════════════════════════════════════════════════════════════
def main():
    """Executa o fluxo RPA completo."""
    inicio = datetime.now()

    print("=" * 70)
    print("[M3S05] Ex. 4 - Fluxo RPA Completo com Arquivos")
    print("=" * 70)

    # ── FASE 1 — Ler planilha de URLs ────────────────────────────────────
    print("\n[1/5] Lendo planilha de URLs...")

    if not PLANILHA_URLS.exists():
        raise FileNotFoundError(f"Planilha de URLs não encontrada: {PLANILHA_URLS}")

    df_urls = pd.read_excel(PLANILHA_URLS, engine="openpyxl")
    print(f"      {len(df_urls)} URLs encontradas:")
    for _, row in df_urls.iterrows():
        print(f"        [{row['id']}] {row['nome']} ({row['tipo']})")
        print(f"            {row['url']}")

    # Validar colunas obrigatórias
    colunas_obrigatorias = {"url", "nome", "tipo"}
    colunas_presentes = set(df_urls.columns)
    faltantes = colunas_obrigatorias - colunas_presentes
    if faltantes:
        raise ValueError(f"Colunas obrigatórias ausentes: {faltantes}")

    # ── FASE 2 — Coletar dados com Selenium ──────────────────────────────
    print("\n[2/5] Iniciando coleta com Selenium...")
    driver = criar_driver()
    data_coleta = datetime.now().isoformat()
    resultados = {}

    try:
        for _, row in df_urls.iterrows():
            nome = row["nome"]
            url = row["url"]
            tipo = row["tipo"]

            print(f"\n      Coletando: {nome}...")

            if tipo not in SCRAPERS:
                print(f"      ⚠ Tipo '{tipo}' não possui scraper. Pulando.")
                continue

            df_dados = SCRAPERS[tipo](driver, url)
            df_dados["fonte"] = nome
            df_dados["data_coleta"] = data_coleta
            df_dados["url_origem"] = url
            resultados[nome] = df_dados

            print(f"      ✓ {len(df_dados)} registros coletados.")

            # Delay aleatório entre requisições
            delay = random.uniform(1, 3)
            print(f"      Aguardando {delay:.1f}s...")
            time.sleep(delay)

    finally:
        driver.quit()
        print("\n      Driver encerrado.")

    # ── FASE 3 — Consolidar dados ────────────────────────────────────────
    print("\n[3/5] Consolidando dados...")
    for nome, df in resultados.items():
        print(f"      {nome}: {df.shape[0]} linhas x {df.shape[1]} colunas")

    # ── FASE 4 — Salvar em planilha formatada ────────────────────────────
    print("\n[4/5] Salvando planilha formatada...")

    DIR_SAIDA.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)  # Remover aba padrão vazia

    # Criar uma aba para cada dataset
    for nome, df in resultados.items():
        # Nome da aba: max 31 chars (limite do Excel)
        nome_aba = nome[:31]
        ws = wb.create_sheet(title=nome_aba)

        # Escrever header
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Escrever dados
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            for col_idx, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

    # Criar aba de metadados
    ws_meta = wb.create_sheet(title="Metadados")
    metadados = [
        ("Campo", "Valor"),
        ("Data/Hora da Coleta", data_coleta),
        ("Total de Fontes", len(resultados)),
        ("Total de Registros", sum(len(df) for df in resultados.values())),
        ("Script", "rpa_full_flow.py"),
        ("Módulo", "M3S05 - DEVinHouse"),
    ]

    # Adicionar info de cada fonte
    for i, (nome, df) in enumerate(resultados.items(), 1):
        metadados.append((f"Fonte {i}", nome))
        metadados.append((f"Fonte {i} - Registros", len(df)))
        metadados.append((f"Fonte {i} - URL", df["url_origem"].iloc[0] if len(df) > 0 else "N/A"))

    for row_idx, (campo, valor) in enumerate(metadados, 1):
        ws_meta.cell(row=row_idx, column=1, value=campo)
        ws_meta.cell(row=row_idx, column=2, value=str(valor))

    # Aplicar formatação profissional
    formatar_planilha(wb)

    wb.save(ARQUIVO_SAIDA_BRUTO)
    print(f"      Planilha salva: {ARQUIVO_SAIDA_BRUTO.name}")

    # ── FASE 5 — Renomear arquivo ────────────────────────────────────────
    print("\n[5/5] Renomeando arquivo de saída...")
    novo_caminho = renomear_arquivo(ARQUIVO_SAIDA_BRUTO)
    print(f"      Arquivo renomeado: {novo_caminho.name}")

    # ── Resumo final ─────────────────────────────────────────────────────
    duracao = (datetime.now() - inicio).total_seconds()
    print("\n" + "=" * 70)
    print("RESUMO DO FLUXO RPA")
    print("=" * 70)
    print(f"  Fontes processadas:   {len(resultados)}")
    print(f"  Total de registros:   {sum(len(df) for df in resultados.values())}")
    print(f"  Arquivo de saída:     {novo_caminho.name}")
    print(f"  Caminho completo:     {novo_caminho}")
    print(f"  Duração total:        {duracao:.1f}s")
    print("=" * 70)
    print("Fluxo RPA concluído com sucesso!")
    print("=" * 70)


if __name__ == "__main__":
    main()
